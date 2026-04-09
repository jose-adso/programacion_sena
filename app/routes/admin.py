from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from flask_login import login_required, current_user
from flask_mail import Message
from openpyxl import load_workbook
from app.models.users import Users
from app import db, mail
from datetime import datetime, timedelta
import smtplib
import random
import string
import re
import xlrd
from urllib.parse import urljoin


def enviar_link_activacion(correo_destino, nombre_usuario, login_username, reset_url):
    """Envía al usuario un enlace para definir su contraseña inicial."""
    msg = Message(
        subject="Activa tu cuenta - Sistema de Competencias SENA",
        recipients=[correo_destino],
        body=f"""
Hola {nombre_usuario},

Se ha creado una cuenta para ti en el Sistema de Competencias SENA.

Tu usuario para iniciar sesion es: {login_username}

Para activar tu acceso, define tu contraseña en el siguiente enlace:
{reset_url}

Este enlace expira en 1 hora.

Si no solicitaste este registro, ignora este correo.

Saludos,
Sistema de Competencias SENA
"""
    )
    try:
        mail.send(msg)
    except smtplib.SMTPAuthenticationError as exc:
        raise RuntimeError(
            "No se pudo enviar el correo por autenticación SMTP. "
            "Verifique el correo remitente y la contraseña de aplicación en Gmail."
        ) from exc
    except Exception as exc:
        raise RuntimeError(
            "No se pudo enviar el correo en este momento."
        ) from exc


def construir_reset_url(user_id, token):
    """Construye la URL pública de activación usando APP_BASE_URL si existe."""
    reset_path = url_for("auth.reset_password", user_id=user_id, token=token)
    app_base_url = (current_app.config.get("APP_BASE_URL") or "").strip()

    if app_base_url:
        return urljoin(app_base_url.rstrip("/") + "/", reset_path.lstrip("/"))

    return url_for("auth.reset_password", user_id=user_id, token=token, _external=True)

def generar_password_aleatoria(longitud=8):
    """Genera una contraseña aleatoria con al menos una mayúscula y un carácter especial"""
    # Garantizar al menos una mayúscula y un carácter especial
    mayusculas = string.ascii_uppercase
    especiales = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    digitos = string.digits
    minusculas = string.ascii_lowercase
    
    # Asegurar que tenga los requisitos
    password = [
        random.choice(mayusculas),  # Al menos una mayúscula
        random.choice(especiales),   # Al menos un carácter especial
        random.choice(digitos),      # Al menos un dígito
    ]
    
    # Llenar el resto con caracteres aleatorios
    todos = mayusculas + especiales + digitos + minusculas
    for _ in range(longitud - 3):
        password.append(random.choice(todos))
    
    # Mezclar la contraseña
    random.shuffle(password)
    return ''.join(password)


def normalize_excel_text(value):
    if value is None:
        return ""

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")

    if isinstance(value, int):
        return str(value)

    return str(value).strip()


def normalize_person_name(value):
    return " ".join(normalize_excel_text(value).split())


def extract_email(value):
    text = normalize_excel_text(value).strip().lower().replace("mailto:", "")
    if not text:
        return ""

    match = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", text, re.IGNORECASE)
    return match.group(0).lower() if match else ""


def get_excel_row_value(row, index):
    if len(row) <= index:
        return None

    cell = row[index]
    if hasattr(cell, "value"):
        hyperlink = getattr(cell, "hyperlink", None)
        if hyperlink and getattr(hyperlink, "target", None):
            return hyperlink.target
        return cell.value

    return cell


def iter_instructor_excel_rows(excel_file, filename):
    if filename.endswith(".xls"):
        file_bytes = excel_file.read()
        excel_file.stream.seek(0)
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        for sheet in workbook.sheets():
            for row_index in range(sheet.nrows):
                yield sheet.name, row_index + 1, sheet.row_values(row_index)
        return

    excel_file.stream.seek(0)
    workbook = load_workbook(excel_file, data_only=False)
    try:
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                yield sheet.title, row_index, row
    finally:
        workbook.close()


def import_instructors_from_excel(excel_file):
    filename = (excel_file.filename or "").lower()
    if not filename.endswith((".xls", ".xlsx")):
        raise ValueError("El archivo debe ser un Excel válido (.xls o .xlsx)")

    existing_users = Users.query.all()
    existing_users_by_email = {
        (user.correo or "").strip().lower(): user
        for user in existing_users
        if user.correo
    }
    existing_users_by_login = {
        user.login_username: user
        for user in existing_users
        if user.correo
    }

    imported_count = 0
    updated_count = 0
    skipped_rows = []
    duplicate_rows = []
    email_error_rows = []
    updated_rows = []

    for sheet_name, row_index, row in iter_instructor_excel_rows(excel_file, filename):
        nombre_raw = get_excel_row_value(row, 2)   # Columna C
        correo_raw = get_excel_row_value(row, 5)    # Columna F
        row_label = f"{sheet_name} fila {row_index}"

        if all(value in (None, "") for value in (nombre_raw, correo_raw)):
            continue

        nombre = normalize_person_name(nombre_raw)
        correo = extract_email(correo_raw)

        if nombre.lower() in {"nombre", "nombres", "instructor", "nombre instructor", "nombre del instructor", "apellidos y nombres", "nombre completo"}:
            continue
        if normalize_excel_text(correo_raw).strip().lower() in {"correo", "correo electronico", "correo electrónico", "email", "e-mail"}:
            continue

        if not nombre or not correo:
            skipped_rows.append(row_label)
            continue

        login_username = correo.split("@", 1)[0]
        existing_user = existing_users_by_email.get(correo)
        login_conflict_user = existing_users_by_login.get(login_username)

        if login_conflict_user and login_conflict_user.correo.strip().lower() != correo:
            duplicate_rows.append(f"{row_label} (usuario '{login_username}' ya existe)")
            continue

        if existing_user and existing_user.is_base_super_admin:
            duplicate_rows.append(f"{row_label} (super admin protegido, no se modifica)")
            continue

        target_user = None

        try:
            with db.session.begin_nested():
                if existing_user:
                    existing_user.nombre = nombre
                    existing_user.rol = "instructor"
                    existing_user.must_change_password = True
                    existing_user.generate_recovery_token()
                    db.session.flush()
                    target_user = existing_user
                    updated_count += 1
                    updated_rows.append(row_label)
                else:
                    password = generar_password_aleatoria(8)
                    new_user = Users(
                        nombre=nombre,
                        correo=correo,
                        telefono="",
                        direccion="",
                        password=password,
                        rol="instructor",
                        must_change_password=True,
                        perfil_profesional=""
                    )
                    db.session.add(new_user)
                    db.session.flush()

                    new_user.generate_recovery_token()
                    db.session.flush()
                    target_user = new_user

                    existing_users_by_email[correo] = new_user
                    existing_users_by_login[login_username] = new_user
                    imported_count += 1
        except Exception:
            skipped_rows.append(row_label)
            continue

        try:
            reset_url = construir_reset_url(target_user.id, target_user.recovery_token)
            enviar_link_activacion(correo, target_user.nombre, target_user.login_username, reset_url)
        except RuntimeError:
            email_error_rows.append(row_label)
        except Exception:
            email_error_rows.append(row_label)

    return imported_count, updated_count, skipped_rows, duplicate_rows, email_error_rows, updated_rows

admin_bp = Blueprint("admin", __name__)

@admin_bp.route("/panel")
@login_required
def panel():
    is_base_super_admin = current_user.is_base_super_admin
    can_open_panel = current_user.rol_activo in ["super admin", "administrador", "gestor"] or is_base_super_admin

    # Super admin base puede entrar al panel para recuperar/cambiar su rol activo.
    if not can_open_panel:
        flash("No tienes permiso para acceder al panel de administrador", "danger")
        return redirect(url_for("main.home"))

    can_manage_users = current_user.rol_activo in ["super admin", "administrador"]
    visible_roles = ["administrador", "gestor", "instructor"]

    if is_base_super_admin:
        visible_roles = ["super admin", *visible_roles]

    if can_manage_users or current_user.rol_activo == "gestor":
        users = Users.query.filter(Users.rol.in_(visible_roles)).all()
    else:
        users = [current_user] if current_user.rol in visible_roles or is_base_super_admin else []

    return render_template(
        "admin/panel.html",
        users=users,
        can_manage_users=can_manage_users,
        is_base_super_admin=is_base_super_admin,
    )


@admin_bp.route("/enviar-links-instructores", methods=["POST"])
@login_required
def enviar_links_instructores():
    """Envía enlace de asignación de contraseña a todos los instructores."""
    if current_user.rol_activo != "super admin":
        flash("Solo el super admin puede enviar enlaces masivos a instructores", "danger")
        return redirect(url_for("admin.panel"))

    instructors = Users.query.filter_by(rol="instructor").all()
    if not instructors:
        flash("No hay instructores registrados para enviar correos", "warning")
        return redirect(url_for("admin.panel"))

    sent = 0
    failed = 0

    for user in instructors:
        try:
            token = user.generate_recovery_token()
            user.must_change_password = True
            db.session.commit()

            reset_url = construir_reset_url(user.id, token)
            enviar_link_activacion(user.correo, user.nombre, user.login_username, reset_url)
            sent += 1
        except RuntimeError:
            db.session.rollback()
            failed += 1
        except Exception:
            db.session.rollback()
            failed += 1

    if sent:
        flash(f"Se enviaron enlaces de asignación de contraseña a {sent} instructores", "success")
    if failed:
        flash(f"No se pudo enviar el correo a {failed} instructores", "warning")

    return redirect(url_for("admin.panel"))

@admin_bp.route("/cambiar_rol/<int:user_id>", methods=["POST"])
@login_required
def cambiar_rol(user_id):
    """Cambiar el rol de un usuario temporalmente"""
    is_base_super_admin = current_user.is_base_super_admin
    is_active_admin = current_user.rol_activo in ["super admin", "administrador"]

    # Super admin base puede cambiar su propio rol aunque su rol activo sea otro.
    if not is_active_admin and not is_base_super_admin:
        return jsonify({"success": False, "error": "No tienes permiso"}), 403

    # Si super admin base está actuando como rol no administrativo, solo puede cambiarse a sí mismo.
    if is_base_super_admin and not is_active_admin and user_id != current_user.id:
        return jsonify({"success": False, "error": "Con tu rol activo actual solo puedes cambiar tu propio rol"}), 403
    
    data = request.get_json()
    nuevo_rol = data.get('rol')
    duracion = data.get('duracion')  # 'semana', 'mes', 'permanente'
    
    user = Users.query.get(user_id)
    if not user:
        return jsonify({"success": False, "error": "Usuario no encontrado"}), 404

    if user.is_base_super_admin and user.id != current_user.id:
        return jsonify({"success": False, "error": "La cuenta base de super admin está protegida"}), 403
    
    # Validar rol
    roles_validos = ["super admin", "administrador", "gestor", "instructor"]
    if nuevo_rol not in roles_validos:
        return jsonify({"success": False, "error": "Rol no válido"}), 400
    
    # Si es permanente o si el usuario actual no es super admin, no permitir crear super admin
    if nuevo_rol == "super admin" and not is_base_super_admin:
        return jsonify({"success": False, "error": "No tienes permiso para asignar rol de super admin"}), 403
    
    hoy = datetime.now().date()

    if user.is_base_super_admin and duracion == 'permanente' and nuevo_rol != 'super admin':
        return jsonify({"success": False, "error": "La cuenta base de super admin no puede perder su rol permanente"}), 403
    
    if duracion == 'permanente':
        # Establecer rol permanente
        user.rol = nuevo_rol
        user.temp_rol = None
        user.temp_rol_start = None
        user.temp_rol_end = None
        mensaje = f"Rol cambiado a {nuevo_rol} permanentemente"
    elif duracion == 'semana':
        # Rol temporal por una semana
        user.temp_rol = nuevo_rol
        user.temp_rol_start = hoy
        user.temp_rol_end = hoy + timedelta(days=7)
        mensaje = f"Rol cambiado a {nuevo_rol} por una semana"
    elif duracion == 'mes':
        # Rol temporal por un mes
        user.temp_rol = nuevo_rol
        user.temp_rol_start = hoy
        user.temp_rol_end = hoy + timedelta(days=30)
        mensaje = f"Rol cambiado a {nuevo_rol} por un mes"
    else:
        return jsonify({"success": False, "error": "Duración no válida"}), 400
    
    try:
        db.session.commit()
        return jsonify({"success": True, "message": mensaje})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route("/registrar", methods=["GET", "POST"])
@login_required
def registrar():
    # Super admin, administrador y gestor pueden ver la pantalla.
    if current_user.rol_activo not in ["super admin", "administrador", "gestor"]:
        flash("No tienes permiso para registrar usuarios", "danger")
        return redirect(url_for("main.home"))
    
    # Gestores solo pueden ver, no registrar
    if request.method == "POST" and current_user.rol_activo == "gestor":
        flash("Los gestores no pueden registrar usuarios", "danger")
        return redirect(url_for("admin.registrar"))
    
    if request.method == "POST":
        submit_type = request.form.get("submit_type", "manual")

        if submit_type == "excel":
            excel_file = request.files.get("excel_file")
            if not excel_file or not excel_file.filename:
                flash("Selecciona un archivo Excel para registrar instructores", "danger")
                return redirect(url_for("admin.registrar"))

            try:
                imported_count, updated_count, skipped_rows, duplicate_rows, email_error_rows, updated_rows = import_instructors_from_excel(excel_file)
                db.session.commit()

                if imported_count or updated_count:
                    flash(
                        f"Carga completada: {imported_count} instructores nuevos y {updated_count} actualizados/reenviados desde Excel.",
                        "success"
                    )
                else:
                    flash("No se registraron instructores. Revisa que el Excel tenga nombres en C y correos válidos en F.", "warning")

                if updated_rows:
                    flash(
                        f"Se actualizaron o reenviaron accesos en: {', '.join(updated_rows[:10])}" + ("..." if len(updated_rows) > 10 else ""),
                        "info"
                    )

                if duplicate_rows:
                    flash(
                        f"No se cargaron filas por correo o usuario ya existente: {', '.join(duplicate_rows[:10])}" + ("..." if len(duplicate_rows) > 10 else ""),
                        "warning"
                    )

                if skipped_rows:
                    flash(
                        f"Se omitieron filas incompletas o inválidas: {', '.join(skipped_rows[:10])}" + ("..." if len(skipped_rows) > 10 else ""),
                        "warning"
                    )

                if email_error_rows:
                    flash(
                        f"Estos instructores sí quedaron registrados/actualizados, pero no se pudo enviar el correo en: {', '.join(email_error_rows[:10])}" + ("..." if len(email_error_rows) > 10 else ""),
                        "warning"
                    )

                return redirect(url_for("admin.panel"))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
                return redirect(url_for("admin.registrar"))
            except Exception:
                db.session.rollback()
                flash("Error al cargar instructores desde Excel", "danger")
                return redirect(url_for("admin.registrar"))

        nombre = (request.form.get("nombre") or "").strip()
        correo = (request.form.get("correo") or "").strip().lower()
        rol = request.form.get("rol")
        perfil_profesional = request.form.get("perfil_profesional", "").strip()
        login_username = correo.split("@", 1)[0] if "@" in correo else ""
        
        # Generar contraseña base aleatoria (no se envía por correo)
        password = generar_password_aleatoria(8)
        
        # Validate role based on current user permissions
        if current_user.rol_activo == "super admin":
            # Super admin can create super admin, administrador, gestor, instructor
            if rol not in ["super admin", "administrador", "gestor", "instructor"]:
                flash("Rol no válido", "danger")
                return redirect(url_for("admin.registrar"))
        elif current_user.rol_activo == "administrador":
            # Administrador can create administrador, gestor, instructor
            if rol not in ["administrador", "gestor", "instructor"]:
                flash("Rol no válido. Solo se pueden registrar administradores, gestores o instructores", "danger")
                return redirect(url_for("admin.registrar"))
        else:
            # Gestor cannot create any users
            flash("No tienes permiso para registrar usuarios", "danger")
            return redirect(url_for("admin.panel"))
        
        if not nombre or not correo or not login_username:
            flash("Debes ingresar un nombre y un correo válido", "danger")
            return redirect(url_for("admin.registrar"))

        existing_user = Users.query.filter(db.func.lower(Users.nombre) == nombre.lower()).first()
        if existing_user:
            flash("Ya existe un usuario con ese nombre", "danger")
            return redirect(url_for("admin.registrar"))

        existing_login_user = Users.query.filter(db.func.lower(Users.correo).like(f"{login_username}@%")).first()
        if existing_login_user:
            flash(f"Ya existe un usuario cuyo acceso es '{login_username}'. Usa otro correo.", "danger")
            return redirect(url_for("admin.registrar"))

        existing_email = Users.query.filter_by(correo=correo).first()
        if existing_email:
            flash("El correo electrónico ya está registrado", "danger")
            return redirect(url_for("admin.registrar"))
        
        # Create new user with only required fields
        new_user = Users(
            nombre=nombre,
            correo=correo,
            telefono="",  # Empty as per requirement
            direccion="",  # Empty as per requirement
            password=password,  # Se hasheará automáticamente gracias al setter
            rol=rol,
            must_change_password=rol in ["administrador", "gestor", "instructor"],
            perfil_profesional=perfil_profesional
        )
        
        try:
            db.session.add(new_user)
            db.session.commit()

            try:
                token = new_user.generate_recovery_token()
                db.session.commit()
                reset_url = construir_reset_url(new_user.id, token)
                enviar_link_activacion(correo, nombre, new_user.login_username, reset_url)
                flash(
                    f"Usuario {nombre} registrado exitosamente como {rol}. Su usuario de acceso es '{new_user.login_username}' y se envió enlace de activación al correo {correo}",
                    "success"
                )
            except RuntimeError as e:
                flash(f"Usuario {nombre} registrado exitosamente como {rol}. {e}", "warning")
            
            return redirect(url_for("admin.panel"))
        except Exception:
            db.session.rollback()
            flash("Error al registrar el usuario", "danger")
            return redirect(url_for("admin.registrar"))
    
    return render_template("admin/registrar.html", read_only=(current_user.rol_activo == "gestor"))

@admin_bp.route("/eliminar/<int:user_id>", methods=["POST"])
@login_required
def eliminar(user_id):
    # Only super admin can delete users
    if current_user.rol_activo not in ["super admin", "administrador"]:
        flash("No tienes permiso para eliminar usuarios", "danger")
        return redirect(url_for("main.home"))
    
    # Gestores solo pueden ver, no eliminar
    if current_user.rol_activo == "gestor":
        flash("Los gestores no pueden eliminar usuarios", "danger")
        return redirect(url_for("admin.panel"))
    
    # Prevent self-deletion
    if user_id == current_user.id:
        flash("No puedes eliminarte a ti mismo", "danger")
        return redirect(url_for("admin.panel"))
    
    user_to_delete = Users.query.get_or_404(user_id)

    if user_to_delete.is_base_super_admin or user_to_delete.rol == "super admin":
        flash("La cuenta base de super admin está protegida y no se puede eliminar", "danger")
        return redirect(url_for("admin.panel"))
    
    try:
        db.session.delete(user_to_delete)
        db.session.commit()
        flash(f"Usuario {user_to_delete.nombre} eliminado exitosamente", "success")
    except Exception:
        db.session.rollback()
        flash("Error al eliminar el usuario", "danger")
    
    return redirect(url_for("admin.panel"))


@admin_bp.route("/password_changes_by_week")
@login_required
def password_changes_by_week():
    if current_user.rol != "super admin":
        return jsonify({"error": "No autorizado"}), 403

    from app.models.users import PasswordHistory
    from calendar import monthrange
    import math

    year = request.args.get('year', type=int, default=datetime.now().year)
    month_0 = request.args.get('month', type=int, default=datetime.now().month - 1)
    month_1 = month_0 + 1  # convertir a 1-based

    days_in_month = monthrange(year, month_1)[1]
    start_dt = datetime(year, month_1, 1)
    end_dt = datetime(year, month_1, days_in_month, 23, 59, 59)

    records = (
        db.session.query(PasswordHistory, Users)
        .join(Users, Users.id == PasswordHistory.user_id)
        .filter(PasswordHistory.created_at >= start_dt,
                PasswordHistory.created_at <= end_dt)
        .order_by(PasswordHistory.created_at.desc())
        .all()
    )

    weeks = {}
    for ph, user in records:
        week_num = math.ceil(ph.created_at.day / 7)
        week_key = f"Semana {week_num}"
        if week_key not in weeks:
            weeks[week_key] = []
        weeks[week_key].append({
            "nombre": user.nombre,
            "rol": user.rol,
            "fecha": ph.created_at.strftime("%d/%m/%Y %H:%M")
        })

    return jsonify({"weeks": weeks, "total": len(records)})