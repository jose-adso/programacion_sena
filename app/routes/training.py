from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from app.models.training import TrainingProgram
from app.models.competency import CompetencyRecord
from app import db
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import re
import unicodedata
import xlrd

training_bp = Blueprint("training", __name__)

WEEK_DAYS = [
    ("lunes", "Lunes"),
    ("martes", "Martes"),
    ("miercoles", "Miercoles"),
    ("jueves", "Jueves"),
    ("viernes", "Viernes"),
    ("sabado", "Sabado"),
    ("domingo", "Domingo"),
]

DEFAULT_EXCEL_COLUMNS = {
    "ficha_number": 4,           # Columna E
    "start_date": 12,            # Columna M
    "end_date": 13,              # Columna N
    "program_name": 26,          # Columna AA
    "location_municipality": 32, # Columna AG
}

REQUIRED_EXCEL_FIELDS = set(DEFAULT_EXCEL_COLUMNS.keys())

EXCEL_HEADER_ALIASES = {
    "ficha_number": {
        "numero de ficha",
        "número de ficha",
        "numero ficha",
        "ficha",
        "nro ficha",
        "no ficha",
        "no de ficha",
        "n ficha",
        "codigo ficha",
        "cod ficha",
    },
    "start_date": {
        "fecha de inicio",
        "fecha inicio",
        "inicio",
    },
    "end_date": {
        "fecha de termino",
        "fecha termino",
        "fecha de terminacion",
        "fecha terminacion",
        "fecha fin",
        "fin",
    },
    "program_name": {
        "nombre del programa",
        "programa de formacion",
        "programa de formación",
        "nombre programa",
        "programa",
    },
    "location_municipality": {
        "lugar o municipio",
        "amb lugar o municipio",
        "municipio",
        "lugar de formacion",
        "lugar de formación",
    },
}

DEFAULT_COMPETENCY_COLUMNS = {
    "program_reference": 2,  # Columna C (la ficha/programa puede venir combinada entre C:F en la fila 6)
    "competencia": 5,        # Columna F
    "resultado": 6,          # Columna G
}

REQUIRED_COMPETENCY_FIELDS = set(DEFAULT_COMPETENCY_COLUMNS.keys())

COMPETENCY_HEADER_ALIASES = {
    "program_reference": {
        "ficha",
        "numero de ficha",
        "número de ficha",
        "nombre de ficha",
        "programa",
        "nombre del programa",
    },
    "competencia": {
        "competencia",
        "competencias",
    },
    "resultado": {
        "resultado",
        "resultado de aprendizaje",
        "resultados de aprendizaje",
        "ra",
    },
}


def normalize_scheduled_days(selected_days):
    day_order = {value: index for index, (value, _) in enumerate(WEEK_DAYS)}
    unique_days = []
    for day in selected_days:
        if day in day_order and day not in unique_days:
            unique_days.append(day)
    return sorted(unique_days, key=lambda day: day_order[day])


def normalize_excel_text(value):
    if value is None:
        return ""

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")

    if isinstance(value, int):
        return str(value)

    return str(value).strip()


def normalize_ficha_number(value):
    ficha_number = normalize_excel_text(value).replace(" ", "")
    if ficha_number.endswith(".0") and ficha_number[:-2].isdigit():
        ficha_number = ficha_number[:-2]
    return ficha_number


def parse_excel_date(value, datemode=None):
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            if datemode is not None:
                return xlrd.xldate_as_datetime(value, datemode).date()
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None

    text_value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue

    return None


def normalize_header_key(value):
    text = normalize_excel_text(value)
    if not text:
        return ""

    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized.lower()).strip()
    return " ".join(normalized.split())


def find_header_row(rows, max_scan_rows=20):
    for row_number, row in enumerate(rows[:max_scan_rows], start=1):
        detected_columns = set()

        for cell_value in row:
            header_key = normalize_header_key(cell_value)
            if not header_key:
                continue

            for field_name, aliases in EXCEL_HEADER_ALIASES.items():
                if field_name in detected_columns:
                    continue

                if any(alias == header_key or alias in header_key for alias in aliases):
                    detected_columns.add(field_name)
                    break

        if REQUIRED_EXCEL_FIELDS.issubset(detected_columns):
            return row_number

    return None


def row_has_valid_program_data(row, datemode):
    ficha_number = normalize_ficha_number(row[DEFAULT_EXCEL_COLUMNS["ficha_number"]] if len(row) > DEFAULT_EXCEL_COLUMNS["ficha_number"] else None)
    program_name = normalize_excel_text(row[DEFAULT_EXCEL_COLUMNS["program_name"]] if len(row) > DEFAULT_EXCEL_COLUMNS["program_name"] else None)
    location_municipality = normalize_excel_text(row[DEFAULT_EXCEL_COLUMNS["location_municipality"]] if len(row) > DEFAULT_EXCEL_COLUMNS["location_municipality"] else None)
    start_date = parse_excel_date(row[DEFAULT_EXCEL_COLUMNS["start_date"]] if len(row) > DEFAULT_EXCEL_COLUMNS["start_date"] else None, datemode)
    end_date = parse_excel_date(row[DEFAULT_EXCEL_COLUMNS["end_date"]] if len(row) > DEFAULT_EXCEL_COLUMNS["end_date"] else None, datemode)

    if ficha_number.lower() in {"número de ficha", "numero de ficha", "ficha", "numeroficha"}:
        return False

    if program_name.lower() in {"nombre del programa", "programa", "programa de formación", "programa de formacion"}:
        return False

    return bool(
        ficha_number
        and program_name
        and location_municipality
        and start_date
        and end_date
        and re.fullmatch(r"\d{5,20}", ficha_number)
    )


def load_excel_rows(excel_file, filename):
    sheets_data = []

    if filename.endswith(".xls"):
        file_bytes = excel_file.read()
        excel_file.stream.seek(0)
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        for sheet in workbook.sheets():
            rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
            sheets_data.append((sheet.name, rows, workbook.datemode))
    else:
        excel_file.stream.seek(0)
        workbook = load_workbook(excel_file, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                sheets_data.append((sheet.title, rows, None))
        finally:
            workbook.close()

    if not sheets_data:
        raise ValueError("El Excel no contiene hojas con datos")

    best_match = None

    for sheet_name, rows, datemode in sheets_data:
        header_row_number = find_header_row(rows)
        valid_rows = 0

        for row_index, row in enumerate(rows, start=1):
            if header_row_number and row_index <= header_row_number:
                continue
            if row_has_valid_program_data(row, datemode):
                valid_rows += 1

        score = (1 if header_row_number else 0, valid_rows)
        if best_match is None or score > best_match[0]:
            best_match = (score, rows, datemode, header_row_number or 0, sheet_name)

    if best_match and best_match[0][1] > 0:
        _, rows, datemode, header_row_number, sheet_name = best_match
        return rows, datemode, header_row_number, DEFAULT_EXCEL_COLUMNS.copy(), sheet_name

    raise ValueError(
        "No se encontró una hoja válida para importar. El Excel debe usar la ficha en E, fechas en M y N, nombre del programa en AA y lugar o municipio en AG."
    )


def find_competency_header_row(rows, max_scan_rows=20):
    for row_number, row in enumerate(rows[:max_scan_rows], start=1):
        detected_columns = set()

        for cell_value in row:
            header_key = normalize_header_key(cell_value)
            if not header_key:
                continue

            for field_name, aliases in COMPETENCY_HEADER_ALIASES.items():
                if field_name in detected_columns:
                    continue

                if any(alias == header_key or alias in header_key for alias in aliases):
                    detected_columns.add(field_name)
                    break

        if REQUIRED_COMPETENCY_FIELDS.issubset(detected_columns):
            return row_number

    return None


def get_next_non_empty_value(row, start_index):
    for value in row[start_index + 1:]:
        text = normalize_excel_text(value)
        if text:
            return text
    return ""


def get_previous_non_empty_value(row, start_index):
    for i in range(start_index - 1, -1, -1):
        text = normalize_excel_text(row[i])
        if text:
            return text
    return ""


def extract_labeled_program_values(row):
    ficha_reference = ""
    program_name = ""

    for index, cell_value in enumerate(row):
        text = normalize_excel_text(cell_value)
        header_key = normalize_header_key(cell_value)
        
        if not text:
            continue

        if not ficha_reference and (
            "ficha de caracterizacion" in header_key or
            text.lower().startswith("ficha de caracterizacion") or
            ("ficha" in header_key and "caracterizacion" in header_key)
        ):
            ficha_reference = get_next_non_empty_value(row, index)
            continue

        if not program_name and (
            "denominacion" in header_key or
            text.lower().startswith("denominacion")
        ):
            program_name = get_next_non_empty_value(row, index)
            continue

        if not program_name and "ficha" in header_key and index > 0:
            prev_value = normalize_excel_text(row[index - 1])
            if prev_value and "denominacion" in normalize_header_key(prev_value):
                program_name = text

        if not ficha_reference and "ficha" in header_key and index > 0:
            prev_value = normalize_excel_text(row[index - 1])
            if prev_value and "ficha de caracterizacion" in normalize_header_key(prev_value):
                ficha_reference = text

    return ficha_reference, program_name


def extract_program_reference_candidates(row):
    candidates = []

    ficha_reference, program_name = extract_labeled_program_values(row)
    for value in (ficha_reference, program_name):
        text = normalize_excel_text(value)
        if text and text not in candidates:
            candidates.append(text)

    for col_index, value in enumerate(row[:10]):
        text = normalize_excel_text(value)
        if not text:
            continue
        
        header_key = normalize_header_key(text)
        
        if "ficha de caracterizacion" in header_key and col_index + 1 < len(row):
            next_val = normalize_excel_text(row[col_index + 1])
            if next_val and next_val not in candidates:
                candidates.append(next_val)
        
        if "denominacion" in header_key and col_index + 1 < len(row):
            next_val = normalize_excel_text(row[col_index + 1])
            if next_val and next_val not in candidates:
                candidates.append(next_val)
        
        if re.match(r"^\d+\s*-", text):
            if text not in candidates:
                candidates.append(text)
            
            if col_index + 1 < len(row):
                next_val = normalize_excel_text(row[col_index + 1])
                if next_val and next_val not in candidates:
                    candidates.append(next_val)

    for value in row[2:6]:
        text = normalize_excel_text(value)
        if text and text not in candidates:
            candidates.append(text)

    joined_text = " ".join(candidates).strip()
    if joined_text and joined_text not in candidates:
        candidates.append(joined_text)

    return candidates


def detect_competency_sheet_layout(rows, programs_by_ficha=None, programs_by_name=None, max_scan_rows=20):
    for row_number, row in enumerate(rows[:max_scan_rows], start=1):
        ficha_reference, program_name = extract_labeled_program_values(row)
        if ficha_reference or program_name:
            return {
                "start_row": 13,
                "program_reference": ficha_reference or program_name,
                "ficha_reference": ficha_reference,
                "program_name": program_name,
            }

        if row_number == 6:
            for candidate in extract_program_reference_candidates(row):
                if not candidate:
                    continue
                if not programs_by_ficha or not programs_by_name:
                    return {
                        "start_row": 13,
                        "program_reference": candidate,
                        "ficha_reference": candidate,
                        "program_name": "",
                    }

                program = resolve_training_program(candidate, programs_by_ficha, programs_by_name)
                if program:
                    return {
                        "start_row": 13,
                        "program_reference": candidate,
                        "ficha_reference": candidate,
                        "program_name": program.program_name,
                    }

    return None


def get_competency_texts_from_row(row):
    competencia = normalize_excel_text(row[DEFAULT_COMPETENCY_COLUMNS["competencia"]] if len(row) > DEFAULT_COMPETENCY_COLUMNS["competencia"] else None)
    resultado = normalize_excel_text(row[DEFAULT_COMPETENCY_COLUMNS["resultado"]] if len(row) > DEFAULT_COMPETENCY_COLUMNS["resultado"] else None)
    return competencia, resultado


def row_has_valid_competency_data(row):
    competencia, resultado = get_competency_texts_from_row(row)

    if competencia.lower() in {"competencia", "competencias"}:
        return False

    if resultado.lower() in {"resultado", "resultado de aprendizaje", "resultados de aprendizaje", "ra"}:
        return False

    return bool(competencia or resultado)


def load_competency_excel_rows(excel_file, filename, programs_by_ficha=None, programs_by_name=None):
    sheets_data = []

    if filename.endswith(".xls"):
        file_bytes = excel_file.read()
        excel_file.stream.seek(0)
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        for sheet in workbook.sheets():
            rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
            sheets_data.append((sheet.name, rows))
    else:
        excel_file.stream.seek(0)
        workbook = load_workbook(excel_file, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                sheets_data.append((sheet.title, rows))
        finally:
            workbook.close()

    if not sheets_data:
        raise ValueError("El Excel no contiene hojas con datos")

    best_match = None
    for sheet_name, rows in sheets_data:
        header_row_number = find_competency_header_row(rows)
        labeled_layout = detect_competency_sheet_layout(rows, programs_by_ficha, programs_by_name)
        start_row_number = max(
            header_row_number or 0,
            labeled_layout["start_row"] if labeled_layout else 0,
        )

        valid_rows = 0
        for row_index, row in enumerate(rows, start=1):
            if start_row_number and row_index <= start_row_number:
                continue
            if row_has_valid_competency_data(row):
                valid_rows += 1

        score = (
            1 if labeled_layout and labeled_layout.get("program_reference") else 0,
            1 if (len(rows) > 13 and row_has_valid_competency_data(rows[13])) else 0,
            valid_rows,
        )
        if best_match is None or score > best_match[0]:
            best_match = (
                score,
                rows,
                start_row_number,
                sheet_name,
                (labeled_layout or {}).get("program_reference", ""),
            )

    if best_match and best_match[0][2] > 0:
        _, rows, start_row_number, sheet_name, initial_program_reference = best_match
        return rows, start_row_number, DEFAULT_COMPETENCY_COLUMNS.copy(), sheet_name, initial_program_reference

    raise ValueError(
        "No se encontró una hoja válida para importar competencias. Verifica que el nombre o número de la ficha/programa esté en la fila 6 entre C:F y que desde la fila 14 la competencia esté en F y el resultado en G."
    )


def resolve_training_program(program_reference, programs_by_ficha, programs_by_name):
    raw_reference = normalize_excel_text(program_reference)

    ficha_key = normalize_ficha_number(raw_reference)
    if ficha_key and ficha_key in programs_by_ficha:
        return programs_by_ficha[ficha_key]

    ficha_match = re.search(r"\d{1,20}", raw_reference)
    if ficha_match:
        matched_ficha = ficha_match.group(0)
        if matched_ficha in programs_by_ficha:
            return programs_by_ficha[matched_ficha]

    name_key = normalize_header_key(raw_reference)
    if name_key and name_key in programs_by_name:
        return programs_by_name[name_key]

    for program_name_key, program in programs_by_name.items():
        if program_name_key and (program_name_key in name_key or name_key in program_name_key):
            return program

    return None


def extract_program_reference_from_competency_row(row):
    for col_index, value in enumerate(row[:10]):
        text = normalize_excel_text(value)
        if not text:
            continue
        
        if re.match(r"^\d+\s*-", text):
            match = re.match(r"^(\d+)\s*-", text)
            if match:
                return match.group(1)
        
        if re.match(r"^\d{5,20}$", text.replace(" ", "")):
            return text
        
        ficha_match = re.search(r"\d{5,20}", text)
        if ficha_match:
            return ficha_match.group(0)
    
    return ""


def find_program_reference_in_row(row, programs_by_ficha, programs_by_name):
    candidates = []
    
    for col_index, value in enumerate(row[:10]):
        text = normalize_excel_text(value)
        if not text:
            continue
        
        if re.match(r"^\d+\s*-", text):
            match = re.match(r"^(\d+)\s*-", text)
            if match:
                candidates.append(match.group(1))
        
        clean_num = text.replace(" ", "")
        if re.match(r"^\d{5,20}$", clean_num):
            candidates.append(clean_num)
        
        ficha_match = re.search(r"\d{5,20}", text)
        if ficha_match and ficha_match.group(0) not in candidates:
            candidates.append(ficha_match.group(0))

    candidates.extend(extract_program_reference_candidates(row))

    for value in row[:10]:
        text = normalize_excel_text(value)
        if text and text not in candidates:
            candidates.append(text)

    joined_text = " ".join(candidates).strip()
    if joined_text and joined_text not in candidates:
        candidates.append(joined_text)

    for candidate in candidates:
        if resolve_training_program(candidate, programs_by_ficha, programs_by_name):
            return candidate

    return ""


def choose_program_reference(explicit_program_reference, detected_reference, current_program_reference, programs_by_ficha, programs_by_name):
    candidates = []

    # Priorizar siempre la ficha detectada en el encabezado del reporte.
    for candidate in (current_program_reference, explicit_program_reference, detected_reference):
        text = normalize_excel_text(candidate)
        if text and text not in candidates:
            candidates.append(text)

    for candidate in candidates:
        if resolve_training_program(candidate, programs_by_ficha, programs_by_name):
            return candidate

    return ""


def import_competencies_from_excel(excel_file):
    filename = (excel_file.filename or "").lower()
    if not filename.endswith((".xls", ".xlsx")):
        raise ValueError("El archivo debe ser un Excel válido (.xls o .xlsx)")

    programs = TrainingProgram.query.all()
    programs_by_ficha = {
        normalize_ficha_number(program.ficha_number): program
        for program in programs
        if program.ficha_number
    }
    programs_by_name = {
        normalize_header_key(program.program_name): program
        for program in programs
        if program.program_name
    }

    existing_records = CompetencyRecord.query.all()
    records_by_program_comp = {
        (record.training_program_id, normalize_header_key(record.competencia)): record
        for record in existing_records
        if record.competencia
    }
    duplicate_keys = {
        (
            record.training_program_id,
            normalize_header_key(record.competencia),
            normalize_header_key(record.resultado or "")
        )
        for record in existing_records
    }

    imported_count = 0
    updated_count = 0
    skipped_rows = []
    unresolved_rows = []
    duplicate_rows = []

    rows, start_row_number, column_map, sheet_name, initial_program_reference = load_competency_excel_rows(
        excel_file,
        filename,
        programs_by_ficha,
        programs_by_name,
    )
    current_program_reference = initial_program_reference or ""

    for row_index, row in enumerate(rows, start=1):
        if start_row_number and row_index <= start_row_number:
            continue

        program_reference_raw = row[column_map["program_reference"]] if len(row) > column_map["program_reference"] else None
        competencia_raw = row[column_map["competencia"]] if len(row) > column_map["competencia"] else None
        resultado_raw = row[column_map["resultado"]] if len(row) > column_map["resultado"] else None

        if all(value in (None, "") for value in (program_reference_raw, competencia_raw, resultado_raw)):
            detected_reference = find_program_reference_in_row(row, programs_by_ficha, programs_by_name)
            if detected_reference:
                current_program_reference = detected_reference
            continue

        competencia, resultado = get_competency_texts_from_row(row)

        detected_reference = find_program_reference_in_row(row, programs_by_ficha, programs_by_name)
        explicit_program_reference = normalize_excel_text(program_reference_raw)
        program_reference = choose_program_reference(
            explicit_program_reference,
            detected_reference,
            current_program_reference,
            programs_by_ficha,
            programs_by_name,
        )

        if detected_reference and not competencia:
            current_program_reference = choose_program_reference(
                "",
                detected_reference,
                current_program_reference,
                programs_by_ficha,
                programs_by_name,
            ) or current_program_reference
            continue

        if program_reference.lower() in {"ficha", "numero de ficha", "número de ficha", "nombre de ficha", "programa", "nombre del programa"}:
            continue
        if competencia.lower() in {"competencia", "competencias"}:
            continue

        if not program_reference or not competencia:
            skipped_rows.append(row_index)
            continue

        current_program_reference = program_reference
        program = resolve_training_program(program_reference, programs_by_ficha, programs_by_name)
        if not program:
            unresolved_rows.append(row_index)
            continue

        duplicate_key = (
            program.id,
            normalize_header_key(competencia),
            normalize_header_key(resultado or "")
        )
        if duplicate_key in duplicate_keys:
            duplicate_rows.append(row_index)
            continue

        new_record = CompetencyRecord(
            training_program_id=program.id,
            competencia=competencia,
            resultado=resultado or None,
            instructor_name="CATALOGO EXCEL",
            horario=None,
        )
        db.session.add(new_record)
        duplicate_keys.add(duplicate_key)
        imported_count += 1

    return imported_count, updated_count, skipped_rows, unresolved_rows, duplicate_rows, sheet_name


def import_programs_from_excel(excel_file):
    filename = (excel_file.filename or "").lower()
    if not filename.endswith((".xls", ".xlsx")):
        raise ValueError("El archivo debe ser un Excel válido (.xls o .xlsx)")

    existing_programs = {
        normalize_ficha_number(program.ficha_number): program
        for program in TrainingProgram.query.all()
        if program.ficha_number is not None
    }
    imported_count = 0
    updated_count = 0
    skipped_rows = []
    updated_rows = []

    rows, datemode, header_row_number, column_map, sheet_name = load_excel_rows(excel_file, filename)

    for row_index, row in enumerate(rows, start=1):
        if header_row_number and row_index <= header_row_number:
            continue

        ficha_number_raw = row[column_map["ficha_number"]] if len(row) > column_map["ficha_number"] else None
        start_date_raw = row[column_map["start_date"]] if len(row) > column_map["start_date"] else None
        end_date_raw = row[column_map["end_date"]] if len(row) > column_map["end_date"] else None
        program_name_raw = row[column_map["program_name"]] if len(row) > column_map["program_name"] else None
        location_raw = row[column_map["location_municipality"]] if len(row) > column_map["location_municipality"] else None

        if all(value in (None, "") for value in (ficha_number_raw, program_name_raw, location_raw, start_date_raw, end_date_raw)):
            continue

        ficha_number = normalize_ficha_number(ficha_number_raw)
        program_name = normalize_excel_text(program_name_raw)
        location_municipality = normalize_excel_text(location_raw)
        start_date = parse_excel_date(start_date_raw, datemode)
        end_date = parse_excel_date(end_date_raw, datemode)

        if ficha_number.lower() in {"número de ficha", "numero de ficha", "ficha", "numeroficha"}:
            continue

        if program_name.lower() in {"nombre del programa", "programa", "programa de formación", "programa de formacion"}:
            continue

        if not ficha_number or not program_name or not location_municipality or not start_date or not end_date:
            skipped_rows.append(row_index)
            continue

        if not re.fullmatch(r"\d{1,20}", ficha_number):
            skipped_rows.append(row_index)
            continue

        existing_program = existing_programs.get(ficha_number)
        if existing_program:
            existing_program.program_name = program_name
            existing_program.classroom = "centro"
            existing_program.location_municipality = location_municipality
            existing_program.start_date = start_date
            existing_program.end_date = end_date
            existing_program.updated_at = datetime.utcnow()
            updated_rows.append(row_index)
            updated_count += 1
            continue

        new_program = TrainingProgram(
            ficha_number=ficha_number,
            program_name=program_name,
            classroom="centro",
            location_municipality=location_municipality,
            start_date=start_date,
            end_date=end_date,
        )
        db.session.add(new_program)
        existing_programs[ficha_number] = new_program
        imported_count += 1

    return imported_count, updated_count, skipped_rows, updated_rows, sheet_name

@training_bp.route("/competencies")
@login_required
def list_competencies():
    if current_user.rol_activo not in ["super admin", "administrador", "gestor", "instructor"]:
        flash("No tienes permiso para acceder al listado de competencias", "danger")
        return redirect(url_for("main.home"))

    competency_records = (
        CompetencyRecord.query
        .join(TrainingProgram, CompetencyRecord.training_program_id == TrainingProgram.id)
        .order_by(TrainingProgram.ficha_number.asc(), CompetencyRecord.competencia.asc(), CompetencyRecord.id.asc())
        .all()
    )

    grouped_competencies = []
    groups_by_program = {}

    for record in competency_records:
        program = record.training_program
        if not program:
            continue

        group = groups_by_program.get(program.id)
        if group is None:
            group = {
                "training_program_id": program.id,
                "ficha_number": program.ficha_number,
                "program_name": program.program_name,
                "records": [],
            }
            groups_by_program[program.id] = group
            grouped_competencies.append(group)

        group["records"].append(record)

    return render_template(
        "training/competencies_list.html",
        grouped_competencies=grouped_competencies,
    )


@training_bp.route("/competencies/delete/<int:program_id>", methods=["POST"])
@login_required
def delete_competencies_group(program_id):
    if current_user.rol_activo not in ["super admin", "administrador"]:
        flash("No tienes permiso para eliminar competencias", "danger")
        return redirect(url_for("training.list_competencies"))

    program = TrainingProgram.query.get_or_404(program_id)

    try:
        deleted_count = CompetencyRecord.query.filter_by(training_program_id=program_id).delete()
        db.session.commit()
        flash(
            f"Se eliminaron {deleted_count} competencias/resultados de la ficha {program.ficha_number} - {program.program_name}",
            "success"
        )
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar las competencias de la ficha {program.ficha_number}: {str(e)}", "danger")

    return redirect(url_for("training.list_competencies"))


@training_bp.route("/competencies/upload", methods=["GET", "POST"])
@login_required
def upload_competencies():
    if current_user.rol_activo not in ["super admin", "administrador", "gestor"]:
        flash("No tienes permiso para acceder a la carga de competencias", "danger")
        return redirect(url_for("main.home"))

    if request.method == "POST" and current_user.rol_activo == "gestor":
        flash("Los gestores no pueden cargar competencias por Excel", "danger")
        return redirect(url_for("training.upload_competencies"))

    if request.method == "POST":
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            flash("Selecciona un archivo Excel para cargar competencias", "danger")
            return redirect(url_for("training.upload_competencies"))

        try:
            imported_count, updated_count, skipped_rows, unresolved_rows, duplicate_rows, sheet_name = import_competencies_from_excel(excel_file)
            db.session.commit()

            if imported_count or updated_count:
                flash(
                    f"Carga completada desde la hoja '{sheet_name}': {imported_count} competencias nuevas y {updated_count} actualizadas.",
                    "success"
                )
            else:
                flash("No se importaron competencias. Verifica el encabezado del reporte: 'Ficha de Caracterización' y 'Denominación' arriba, y desde la fila 14 la competencia en F y el resultado en G.", "warning")

            if unresolved_rows:
                flash(
                    f"No se encontraron fichas/programas para las filas: {', '.join(map(str, unresolved_rows[:15]))}" + ("..." if len(unresolved_rows) > 15 else ""),
                    "warning"
                )

            if duplicate_rows:
                flash(
                    f"Se omitieron competencias repetidas en las filas: {', '.join(map(str, duplicate_rows[:15]))}" + ("..." if len(duplicate_rows) > 15 else ""),
                    "info"
                )

            if skipped_rows:
                flash(
                    f"Se omitieron filas incompletas o inválidas: {', '.join(map(str, skipped_rows[:15]))}" + ("..." if len(skipped_rows) > 15 else ""),
                    "warning"
                )

            return redirect(url_for("training.upload_competencies"))
        except ValueError as e:
            db.session.rollback()
            flash(str(e), "danger")
            return redirect(url_for("training.upload_competencies"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al cargar competencias desde Excel: {str(e)}", "danger")
            return redirect(url_for("training.upload_competencies"))

    return render_template(
        "training/upload_competencies.html",
        read_only=(current_user.rol_activo == "gestor"),
    )


@training_bp.route("/programs")
@login_required
def programs():
    # Super admin, administrador y gestor pueden ver programas.
    if current_user.rol_activo not in ["super admin", "administrador", "gestor"]:
        flash("No tienes permiso para acceder a la gestión de programas de formación", "danger")
        return redirect(url_for("main.home"))
    
    programs = TrainingProgram.query.all()
    return render_template("training/programs.html", programs=programs)

@training_bp.route("/programs/add", methods=["GET", "POST"])
@login_required
def add_program():
    # Super admin, administrador y gestor pueden ver la pantalla.
    if current_user.rol_activo not in ["super admin", "administrador", "gestor"]:
        flash("No tienes permiso para agregar programas de formación", "danger")
        return redirect(url_for("main.home"))
    
    # Gestores solo pueden ver, no modificar
    if request.method == "POST" and current_user.rol_activo == "gestor":
        flash("Los gestores no pueden agregar programas", "danger")
        return redirect(url_for("training.programs"))
    
    if request.method == "POST":
        submit_type = request.form.get("submit_type", "manual")

        if submit_type == "excel":
            excel_file = request.files.get("excel_file")
            if not excel_file or not excel_file.filename:
                flash("Selecciona un archivo de Excel para cargar las fichas", "danger")
                return redirect(url_for("training.add_program"))

            try:
                imported_count, updated_count, skipped_rows, updated_rows, sheet_name = import_programs_from_excel(excel_file)
                if imported_count == 0 and updated_count == 0:
                    flash("No se importaron fichas. Revisa que el Excel tenga encabezados válidos y datos completos por fila.", "warning")
                else:
                    db.session.commit()
                    flash(
                        f"Carga completada desde la hoja '{sheet_name}' usando E, M, N, AA y AG: {imported_count} nuevas y {updated_count} actualizadas.",
                        "success"
                    )

                if updated_rows:
                    flash(f"Se actualizaron fichas existentes en las filas: {', '.join(map(str, updated_rows[:15]))}" + ("..." if len(updated_rows) > 15 else ""), "info")

                if skipped_rows:
                    flash(f"Se omitieron filas incompletas o inválidas: {', '.join(map(str, skipped_rows[:15]))}" + ("..." if len(skipped_rows) > 15 else ""), "warning")
                return redirect(url_for("training.programs"))
            except ValueError as e:
                db.session.rollback()
                flash(str(e), "danger")
                return redirect(url_for("training.add_program"))
            except Exception as e:
                db.session.rollback()
                flash(f"Error al cargar el Excel de fichas: {str(e)}", "danger")
                return redirect(url_for("training.add_program"))

        ficha_number = request.form.get("ficha_number")
        program_name = request.form.get("program_name")
        classroom = request.form.get("classroom")
        location_municipality = request.form.get("location_municipality")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        scheduled_days = normalize_scheduled_days(request.form.getlist("scheduled_days"))
        
        # Validaciones básicas
        if not ficha_number or not program_name or not start_date or not end_date:
            flash("Todos los campos marcados como obligatorios deben ser completados", "danger")
            return redirect(url_for("training.add_program"))
        
        # Verificar si la ficha ya existe
        existing_ficha = TrainingProgram.query.filter_by(ficha_number=ficha_number).first()
        if existing_ficha:
            flash("Ya existe un programa con ese número de ficha", "danger")
            return redirect(url_for("training.add_program"))
        
        try:
            # Convertir strings de fecha a objetos date
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            new_program = TrainingProgram(
                ficha_number=ficha_number,
                program_name=program_name,
                classroom=classroom if classroom else None,
                location_municipality=location_municipality if location_municipality else None,
                start_date=start_date_obj,
                end_date=end_date_obj
            )
            
            db.session.add(new_program)
            db.session.commit()
            flash(f"Programa de formación '{program_name}' agregado exitosamente", "success")
            return redirect(url_for("training.programs"))
        except ValueError:
            flash("Formato de fecha inválido. Use YYYY-MM-DD", "danger")
            return redirect(url_for("training.add_program"))
        except Exception:
            db.session.rollback()
            flash("Error al agregar el programa de formación", "danger")
            return redirect(url_for("training.add_program"))
    
    return render_template(
        "training/add_program.html",
        read_only=(current_user.rol_activo == "gestor"),
        week_days=WEEK_DAYS,
    )

@training_bp.route("/programs/edit/<int:program_id>", methods=["GET", "POST"])
@login_required
def edit_program(program_id):
    # Super admin, administrador y gestor pueden ver la pantalla.
    if current_user.rol_activo not in ["super admin", "administrador", "gestor"]:
        flash("No tienes permiso para editar programas de formación", "danger")
        return redirect(url_for("main.home"))
    
    # Gestores solo pueden ver, no modificar
    if request.method == "POST" and current_user.rol_activo == "gestor":
        flash("Los gestores no pueden editar programas", "danger")
        return redirect(url_for("training.programs"))
    
    program = TrainingProgram.query.get_or_404(program_id)
    
    if request.method == "POST":
        ficha_number = request.form.get("ficha_number")
        program_name = request.form.get("program_name")
        classroom = request.form.get("classroom")
        location_municipality = request.form.get("location_municipality")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        scheduled_days = normalize_scheduled_days(request.form.getlist("scheduled_days"))
        
        # Validaciones básicas
        if not ficha_number or not program_name or not start_date or not end_date:
            flash("Todos los campos marcados como obligatorios deben ser completados", "danger")
            return redirect(url_for("training.edit_program", program_id=program_id))

        if not scheduled_days:
            flash("Debes seleccionar al menos un dia de formacion", "danger")
            return redirect(url_for("training.edit_program", program_id=program_id))
        
        # Verificar si la ficha ya existe en otro programa
        existing_ficha = TrainingProgram.query.filter_by(ficha_number=ficha_number).first()
        if existing_ficha and existing_ficha.id != program_id:
            flash("Ya existe otro programa con ese número de ficha", "danger")
            return redirect(url_for("training.edit_program", program_id=program_id))
        
        try:
            # Convertir strings de fecha a objetos date
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            program.ficha_number = ficha_number
            program.program_name = program_name
            program.classroom = classroom if classroom else None
            program.location_municipality = location_municipality if location_municipality else None
            program.start_date = start_date_obj
            program.end_date = end_date_obj
            program.set_scheduled_days(scheduled_days)
            program.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash(f"Programa de formación '{program_name}' actualizado exitosamente", "success")
            return redirect(url_for("training.programs"))
        except ValueError as e:
            flash("Formato de fecha inválido. Use YYYY-MM-DD", "danger")
            return redirect(url_for("training.edit_program", program_id=program_id))
        except Exception as e:
            db.session.rollback()
            flash("Error al actualizar el programa de formación", "danger")
            return redirect(url_for("training.edit_program", program_id=program_id))
    
    # Para GET request, mostrar el formulario con los datos actuales
    return render_template(
        "training/edit_program.html",
        program=program,
        read_only=(current_user.rol_activo == "gestor"),
        week_days=WEEK_DAYS,
    )

@training_bp.route("/programs/delete/<int:program_id>", methods=["POST"])
@login_required
def delete_program(program_id):
    # Only super admin can delete training programs
    if current_user.rol_activo not in ["super admin", "administrador"]:
        flash("No tienes permiso para eliminar programas de formación", "danger")
        return redirect(url_for("main.home"))
    
    # Gestores solo pueden ver, no eliminar
    if current_user.rol_activo == "gestor":
        flash("Los gestores no pueden eliminar programas", "danger")
        return redirect(url_for("training.programs"))
    
    program = TrainingProgram.query.get_or_404(program_id)
    
    try:
        # Eliminar primero los registros relacionados en otras tablas
        from app.models.competency import CompetencyRecord, CalendarAssignment
        
        # Eliminar CalendarAssignments relacionados
        CalendarAssignment.query.filter_by(training_program_id=program_id).delete()
        
        # Eliminar CompetencyRecords relacionados
        CompetencyRecord.query.filter_by(training_program_id=program_id).delete()
        
        # Ahora eliminar el programa
        db.session.delete(program)
        db.session.commit()
        flash(f"Programa de formación '{program.program_name}' eliminado exitosamente", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar el programa de formación: {str(e)}", "danger")
    
    return redirect(url_for("training.programs"))